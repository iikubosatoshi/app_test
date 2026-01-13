import io
import csv
import tempfile
from typing import Dict, List, Any, Tuple

import streamlit as st
import pandas as pd
import openpyxl as excel
from openpyxl.workbook import Workbook


# =========================
# Constants
# =========================
DEFAULT_MAPPING_SHEET = "PowerBoard"   # 研究室対応表のシート名
DEFAULT_OUTPUT_DATA_SHEET = "PowerConsum"
TOTAL_SHEET_NAME = "TotalPower"


# =========================
# Core functions
# =========================
def read_room_to_lab_from_xlsx(xlsx_bytes: bytes, sheet_name: str) -> Dict[Any, Any]:
    """研究室分類リストファイルを読み込み、辞書データ（部屋番号：所属研究室）を返す"""
    bio = io.BytesIO(xlsx_bytes)
    book = excel.load_workbook(bio, data_only=True)
    if sheet_name not in book.sheetnames:
        raise ValueError(f"対応表のシート '{sheet_name}' が見つかりません。存在: {book.sheetnames}")
    sheet = book[sheet_name]

    # 元コード同様：2行目が部屋番号、3行目が研究室名（列方向に並んでいる前提）
    room_to_lab: Dict[Any, Any] = {}
    col_count = 1
    for _col in sheet.iter_cols(min_col=2):
        col_count += 1
        room = sheet.cell(2, col_count).value
        lab = sheet.cell(3, col_count).value
        if room is not None:
            room_to_lab[room] = lab

    if not room_to_lab:
        raise ValueError("対応表から部屋→研究室のデータが取得できませんでした。（2行目/3行目の構造を確認してください）")

    return room_to_lab


def read_power_csv_to_rows(csv_bytes: bytes, encoding: str) -> List[List[Any]]:
    """
    先頭に空行（カンマだけの行）やメタ行があっても壊れない読み取り。
    想定フォーマット（例）:
      - 空行（,,,,）
      - メタ行（log_xxx.csv, ...）
      - ラベル行（先頭空で、"1LP-DK-2 電灯" など）
      - 単位行（年月,kWh,kWh,...）  ← これを基準にする
      - データ行（2025年1月, ...）
    戻り値は「ラベル行」から下（ラベル行+単位行+データ行）を返す。
    """
    text = csv_bytes.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))

    all_rows = []
    for row in reader:
        # 行全体が空（カンマだけ等）なら捨てる
        if not row or all((c is None) or (str(c).strip() == "") for c in row):
            continue
        all_rows.append(row)

    if not all_rows:
        return []

    # 「年月」行を探す（これが単位行の先頭列）
    ym_idx = None
    for i, row in enumerate(all_rows):
        if len(row) >= 1 and str(row[0]).strip() == "年月":
            ym_idx = i
            break

    # 「年月」行が見つかった場合：その1つ上がラベル行（部屋番号/名称が並ぶ行）である前提
    if ym_idx is not None:
        start = max(ym_idx - 1, 0)  # ラベル行
        return all_rows[start:]

    # 見つからない場合のフォールバック：
    # 先頭が "log_" を含むメタ行を1行だけ飛ばす（あるなら）
    if "log_" in str(all_rows[0][0]):
        return all_rows[1:]

    return all_rows

def write_rows_to_workbook_sheet_as_in_original(
    wb: Workbook,
    rows: List[List[Any]],
    sheet_title: str,
) -> None:
    """
    元コードの csv_excel() の書き込みルールに合わせて、PowerConsum シートへ書き込む。
    - 2列目以降 & 3行目以降 & データあり -> float で書き込み
    - それ以外 -> 文字列として書き込み（元の挙動）
    """
    if sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
        wb.remove(ws)

    ws = wb.create_sheet(title=sheet_title)

    cnt_r = 1
    for row in rows:
        cnt_c = 1
        for col in row:
            if col:  # 空でなければ
                if cnt_c > 1 and cnt_r > 2:
                    # 数値化して書き込み（失敗したら文字のまま）
                    try:
                        ws.cell(row=cnt_r, column=cnt_c, value=float(col))
                    except Exception:
                        ws.cell(row=cnt_r, column=cnt_c, value=col)
                else:
                    ws.cell(row=cnt_r, column=cnt_c, value=col)
            else:
                ws.cell(row=cnt_r, column=cnt_c, value=col)
            cnt_c += 1
        cnt_r += 1


def delete_water_columns_like_v14(ws) -> None:
    """
    元コード Ver14 相当の列削除を反映。
    - 109..136 列（= range(109,137)）削除
    - 4..5 列（= range(4,6)）削除
    """
    delete_cols_1 = list(range(109, 137))
    for col in sorted(delete_cols_1, reverse=True):
        ws.delete_cols(col)

    delete_cols_2 = list(range(4, 6))
    for col in sorted(delete_cols_2, reverse=True):
        ws.delete_cols(col)


def split_into_lab_sheets(
    wb: Workbook,
    power_sheet_name: str,
    room_to_lab: Dict[Any, Any],
) -> Tuple[int, int]:
    """
    PowerConsum シートの1行目（部屋番号列）を見て、研究室ごとにシートを作成/追記する。
    返り値: (処理できた部屋数, 対応表に見つからずスキップした部屋数)
    """
    if power_sheet_name not in wb.sheetnames:
        raise ValueError(f"'{power_sheet_name}' シートが見つかりません。")

    src = wb[power_sheet_name]

    processed = 0
    skipped = 0

    # 1行目の2列目以降が部屋番号の想定
    col_n = 1
    for _ in src.iter_cols(min_col=2, min_row=1, max_row=1):
        col_n += 1
        present_room = src.cell(1, col_n).value
        if present_room is None:
            continue

        if present_room not in room_to_lab:
            skipped += 1
            continue

        sh_name = room_to_lab[present_room]
        sh_name_str = str(sh_name)

        if sh_name_str not in wb.sheetnames:
            # 新規シート作成：日時列(1列目) + 該当部屋列(col_n) をコピー
            to_sheet = wb.create_sheet(title=sh_name_str)
            mycols = [1, col_n]

            j = 1
            for cidx in mycols:
                i = 1
                for row in src.iter_rows():
                    cell = row[cidx - 1]  # 0-index
                    if cell.value is not None:
                        if j == 2 and i > 2:
                            try:
                                to_sheet.cell(row=i, column=j).value = float(cell.value)
                            except Exception:
                                to_sheet.cell(row=i, column=j).value = cell.value
                        else:
                            to_sheet.cell(row=i, column=j).value = cell.value
                    i += 1
                j += 1
        else:
            # 既存シートへ追記：最大列+1へ該当部屋列を追加
            to_sheet = wb[sh_name_str]
            maxcol = to_sheet.max_column
            i = 1
            for row in src.iter_rows(min_col=col_n, max_col=col_n):
                cell = row[0]
                if cell.value is not None:
                    if i > 2:
                        try:
                            to_sheet.cell(row=i, column=maxcol + 1).value = float(cell.value)
                        except Exception:
                            to_sheet.cell(row=i, column=maxcol + 1).value = cell.value
                    else:
                        to_sheet.cell(row=i, column=maxcol + 1).value = cell.value
                i += 1

        processed += 1

    return processed, skipped


def calc_total_power_and_cost(
    wb: Workbook,
    price_yen_per_kwh: float,
    target_sheet_names: List[str] | None = None,
) -> pd.DataFrame:
    exclude = {DEFAULT_OUTPUT_DATA_SHEET, TOTAL_SHEET_NAME, "Sheet"}
    candidate = [s for s in wb.sheetnames if s not in exclude]

    if target_sheet_names is None:
        targets = candidate
    else:
        targets = [s for s in target_sheet_names if s in wb.sheetnames and s not in exclude]

    labs = []
    powers = []
    power_units = []
    costs = []
    cost_units = []

    for sh in targets:
        ws = wb[sh]
        power = 0.0

        for col in ws.iter_cols(min_col=2, min_row=3):
            for cell in col:
                if cell.value is None:
                    continue
                try:
                    power += float(cell.value)
                except Exception:
                    continue

        labs.append(sh)
        powers.append(power)
        power_units.append("kWh")
        costs.append(power * float(price_yen_per_kwh))
        cost_units.append("Yen")

    # ===== Excel: 横書き TotalPower =====
    if TOTAL_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[TOTAL_SHEET_NAME])
    ws_total = wb.create_sheet(title=TOTAL_SHEET_NAME)

    ws_total.append(labs)
    ws_total.append(power_units)
    ws_total.append(powers)
    ws_total.append(cost_units)
    ws_total.append(costs)

    # ===== UI表示用 DataFrame（縦） =====
    df = pd.DataFrame({
        "研究室": labs,
        "使用量(kWh)": powers,
        "料金(円)": costs,
    }).sort_values("料金(円)", ascending=False).reset_index(drop=True)

    return df
    
def build_output_excel_bytes(
    mapping_xlsx_bytes: bytes,
    power_csv_bytes: bytes,
    mapping_sheet_name: str,
    csv_encoding: str,
    price_yen_per_kwh: float,
) -> Tuple[bytes, pd.DataFrame, Dict[str, Any]]:
    """
    アップロードされたファイルから、出力Excel（bytes）と集計DFを作る。
    """
    room_to_lab = read_room_to_lab_from_xlsx(mapping_xlsx_bytes, mapping_sheet_name)
    rows = read_power_csv_to_rows(power_csv_bytes, csv_encoding)

    if not rows:
        raise ValueError("電気CSVが空か、ヘッダしかありません。")

    wb = excel.Workbook()
    # openpyxl は最初に "Sheet" ができるので消しておく（管理しやすく）
    default_ws = wb.active
    wb.remove(default_ws)

    write_rows_to_workbook_sheet_as_in_original(wb, rows, DEFAULT_OUTPUT_DATA_SHEET)
    
    # 取り込んだ PowerConsum のヘッダ正規化：
    # 1行目が空なら、2行目/3行目のどちらかをヘッダ扱いにする（ケース対応）
    ws = wb[DEFAULT_OUTPUT_DATA_SHEET]
    if ws.cell(1, 2).value is None:
        # 2行目がメタ（log_...）なら3行目をヘッダに寄せる、など
        # ここでは「1行目が空」なら「2行目を1行目へ繰り上げ」する例
        ws.delete_rows(1)

    # 水道列削除（Ver14）
    delete_water_columns_like_v14(wb[DEFAULT_OUTPUT_DATA_SHEET])

    processed, skipped = split_into_lab_sheets(wb, DEFAULT_OUTPUT_DATA_SHEET, room_to_lab)

    df_total = calc_total_power_and_cost(wb, price_yen_per_kwh)

    # 保存してbytes化
    out = io.BytesIO()
    wb.save(out)
    out_bytes = out.getvalue()

    meta = {
        "rooms_in_mapping": len(room_to_lab),
        "rooms_processed": processed,
        "rooms_skipped_not_in_mapping": skipped,
        "lab_sheets_count": len([s for s in wb.sheetnames if s not in {DEFAULT_OUTPUT_DATA_SHEET, TOTAL_SHEET_NAME}]),
    }
    return out_bytes, df_total, meta


# =========================
# UI (Streamlit)
# =========================
st.set_page_config(page_title="D棟 電気使用量 集計", layout="wide")
st.title("D棟 電気使用量 集計アプリ（Streamlit版）")

with st.sidebar:
    st.header("入力")
    mapping_sheet_name = st.text_input("対応表シート名", value=DEFAULT_MAPPING_SHEET)
    csv_encoding = st.selectbox("電気CSVエンコーディング", ["cp932", "utf-8", "utf-8-sig"], index=0)
    price = st.number_input("単価（円/kWh）", min_value=0.0, value=30.0, step=0.5)

    st.caption("※ 対応表：2行目=部屋番号、3行目=研究室名（列方向）を想定")

col1, col2 = st.columns(2)

with col1:
    mapping_file = st.file_uploader("研究室対応表（例：D棟管理区域表.xlsx）", type=["xlsx"], key="mapping")
with col2:
    power_csv = st.file_uploader("電気使用量CSV（log_....csv）", type=["csv"], key="powercsv")

run = st.button("集計する", type="primary", use_container_width=True)

if run:
    if not mapping_file or not power_csv:
        st.error("対応表（xlsx）と電気CSV（csv）の両方をアップロードしてください。")
        st.stop()

    try:
        out_bytes, df_total, meta = build_output_excel_bytes(
            mapping_xlsx_bytes=mapping_file.getvalue(),
            power_csv_bytes=power_csv.getvalue(),
            mapping_sheet_name=mapping_sheet_name,
            csv_encoding=csv_encoding,
            price_yen_per_kwh=price,
        )

        st.success("集計が完了しました。")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("対応表の部屋数", meta["rooms_in_mapping"])
        m2.metric("処理した部屋数", meta["rooms_processed"])
        m3.metric("対応表に無くスキップ", meta["rooms_skipped_not_in_mapping"])
        m4.metric("研究室シート数", meta["lab_sheets_count"])

        st.subheader("研究室別 集計（TotalPower）")
        st.dataframe(df_total, use_container_width=True)

        total_kwh = float(df_total["使用量(kWh)"].sum()) if not df_total.empty else 0.0
        total_yen = float(df_total["料金(円)"].sum()) if not df_total.empty else 0.0
        c1, c2 = st.columns(2)
        c1.metric("総使用量(kWh)", f"{total_kwh:,.2f}")
        c2.metric("総料金(円)", f"{total_yen:,.0f}")

        st.download_button(
            label="Excelをダウンロード（研究室別シート＋TotalPower）",
            data=out_bytes,
            file_name="P_Consum_at_D.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        if meta["rooms_skipped_not_in_mapping"] > 0:
            st.warning("対応表に存在しない部屋番号がCSV側に含まれており、一部スキップされています。対応表を更新すると解消できます。")

    except Exception as e:
        st.exception(e)

st.divider()
st.caption("元スクリプトの処理フロー（CSV→Excel化、水道列削除、研究室別シート作成/追記、合計・料金計算）をStreamlit化しています。")